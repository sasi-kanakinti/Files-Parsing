import os
from io import StringIO
import pandas as pd
from openpyxl import load_workbook
from typing import Tuple, List


def parse_excel(file_path: str, session_id: str) -> Tuple[str, List[str]]:
   
    saved_images = []

    images_dir = os.path.join("/tmp/outputs/images", "excel_images", session_id)
    os.makedirs(images_dir, exist_ok=True)

    df = pd.read_excel(file_path, engine="openpyxl")
    buffer = StringIO()
    df.to_csv(buffer, index=False)
    csv_content = buffer.getvalue()

    try:
        wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            for idx, image in enumerate(getattr(ws, "_images", []) or []):
                img_name = f"{os.path.basename(file_path)}_{sheet_name}_{idx+1}.png"
                img_path = os.path.join(images_dir, img_name)

                try:
                    image.image.save(img_path)
                    saved_images.append(img_path)
                except Exception:
                    try:
                        with open(img_path, "wb") as fh:
                            fh.write(image.ref)
                        saved_images.append(img_path)
                    except:
                        pass

    except Exception:
        pass

    return csv_content, saved_images
